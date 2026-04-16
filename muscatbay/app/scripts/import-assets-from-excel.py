#!/usr/bin/env python3
import json
import os
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests


WORKBOOK_PATH = Path("/Users/sam24/Downloads/muscatbay_app/Muscat_Bay_Asset_Register_Enhanced.xlsx")
SHEET_NAME = "Master_Asset_Register"
TABLE_NAME = "Assets_Register_Database"
BATCH_SIZE = 200


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def normalize(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        lower = stripped.lower()
        if lower == "true":
            return True
        if lower == "false":
            return False
        return stripped
    return value


def main() -> None:
    env_path = Path(__file__).resolve().parents[1] / ".env.local"
    load_env(env_path)

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")

    if not WORKBOOK_PATH.exists():
        raise RuntimeError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in ws[1]]
    if any(h is None for h in headers):
        raise RuntimeError("Header row contains empty columns.")
    headers = [str(h).strip() for h in headers]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        record = {headers[i]: normalize(row[i]) for i in range(len(headers))}
        rows.append(record)

    endpoint = f"{supabase_url}/rest/v1/{TABLE_NAME}"
    headers_http = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    uploaded = 0
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        response = requests.post(endpoint, headers=headers_http, data=json.dumps(batch), timeout=90)
        if response.status_code >= 300:
            raise RuntimeError(
                f"Insert failed for batch {i // BATCH_SIZE + 1}: {response.status_code} {response.text[:1000]}"
            )
        uploaded += len(batch)
        print(f"Uploaded {uploaded}/{len(rows)}")

    count_resp = requests.get(
        endpoint,
        headers={**headers_http, "Prefer": "count=exact"},
        params={"select": "Asset_UID", "limit": 1},
        timeout=30,
    )
    count_resp.raise_for_status()
    count_header = count_resp.headers.get("content-range", "*/0")
    total = count_header.split("/")[-1]
    print(f"Done. Sheet rows imported: {len(rows)} | DB row count: {total}")


if __name__ == "__main__":
    main()
