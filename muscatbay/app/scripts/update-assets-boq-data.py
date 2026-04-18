#!/usr/bin/env python3
"""
Populate BOQ and Reserve Fund columns in Assets_Register_Database from the
enriched Excel workbook.  Matches rows by Asset_UID and PATCHes only the
four new columns; existing data is left untouched.

Usage:
    python3 scripts/update-assets-boq-data.py
"""

import json
import os
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests

WORKBOOK_PATH = Path("/Users/sam24/Downloads/muscatbay_app/Assets_Register_Tracker_v2_Reserve_Fund_Enriched.xlsx")
SHEET_NAME    = "Master_Asset_Register"
TABLE_NAME    = "Assets_Register_Database"
BATCH_SIZE    = 200

BOQ_COLUMNS = [
    "BOQ_Project_Ref",
    "BOQ_Category_Design_Life",
    "BOQ_Unit_Cost_OMR",
    "Current_Replacement_Cost_OMR",
]


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def normalize(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return value


def main() -> None:
    env_path = Path(__file__).resolve().parents[1] / ".env.local"
    load_env(env_path)

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key  = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")

    if not WORKBOOK_PATH.exists():
        raise RuntimeError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    raw_headers = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    uid_idx = headers.index("Asset_UID")
    boq_indices = {col: headers.index(col) for col in BOQ_COLUMNS if col in headers}

    missing = set(BOQ_COLUMNS) - set(boq_indices)
    if missing:
        print(f"Warning: columns not found in Excel and will be skipped: {missing}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        uid = normalize(row[uid_idx])
        if not uid:
            continue
        patch = {"Asset_UID": uid}
        for col, idx in boq_indices.items():
            patch[col] = normalize(row[idx])
        rows.append(patch)

    print(f"Read {len(rows)} rows from Excel. Sending PATCH batches …")

    endpoint = f"{supabase_url}/rest/v1/{TABLE_NAME}"
    http_headers = {
        "apikey":        service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type":  "application/json",
        "Prefer":        "return=minimal,resolution=merge-duplicates",
    }

    updated = 0
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        resp = requests.post(
            endpoint,
            headers={**http_headers, "Prefer": "return=minimal,resolution=merge-duplicates"},
            params={"on_conflict": "Asset_UID"},
            data=json.dumps(batch),
            timeout=90,
        )
        if resp.status_code >= 300:
            raise RuntimeError(
                f"Upsert failed for batch starting at row {i}: "
                f"{resp.status_code} {resp.text[:500]}"
            )
        updated += len(batch)
        print(f"  Updated {updated}/{len(rows)} rows …")

    print(f"\nDone. {updated} rows upserted with BOQ data.")


if __name__ == "__main__":
    main()
